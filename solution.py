import sys
import subprocess
import openpyxl
from openpyxl import load_workbook

#this appliction will extract data from PDF files and populate an excel sheet

#extract all data from PDF files
#PDF --> .txt
def pdfToTxt(pdf_list):
    #will hold the txt file names
    txt_list = []
    # commands
    cmd = "pdf2txt.py"
    op1 = "-o"
    op2 = "-t"
    op2Specifier ="text"

    for pdf in pdf_list:
        name = pdf.split(".")
        textFile = name[0] + ".txt"
        txt_list.append(textFile)
        #execute pdt2txt
        subprocess.call([cmd, op1, textFile, op2, op2Specifier, pdf])

    #at this point all data has been extracted from the pdf files into the txt files
    #send txt_list to populateEx
    populateEx(txt_list)

#extract specific data from text files
#append to excel sheet
def populateEx(txt_list):
    excel = raw_input("input excel file name: ")
    # open excel file for appending
    eFile = openpyxl.load_workbook(excel)
    sheet = eFile.worksheets[0]
    sDateRow = 2
    eDateRow = 2
    productNameRow = 2
    stateRow = 2
    graRow = 2

    for file in txt_list:
        f = open(file)
        for line in f:
            strings = line.split(" ")
            if strings[0] == "Valid": #then this line contains the start and end date
                startDate = strings[4]
                endDate = strings[6]
                sheet.cell(sDateRow,1).value = startDate
                sheet.cell(eDateRow,2).value = endDate
                sDateRow+=1
                eDateRow+=1

            if len(strings) > 1:
                if strings[1] == "Name:":  # this line contains product name
                    productName = ""
                    state = strings[2]
                    for i in range(2, len(strings)):
                        productName = productName + strings[i] + " "
                    sheet.cell(productNameRow,3).value = productName
                    sheet.cell(stateRow, 4).value = state
                    stateRow+=1
                    productNameRow+=1

                if strings[1] == "*\n": #this line contains the group rating areas
                    gra = strings[0]
                    gra = gra[len(gra) -2] + gra[len(gra) - 1]
                    sheet.cell(graRow, 5).value = gra
                    graRow+=1
    eFile.save(excel)






if __name__ == '__main__':

    # list to hold pdf file names
    pdf_list = []

    #getting command line arguments and populating pdf_list
    if len(sys.argv) >= 2:
        for i in range(1,len(sys.argv)):
            pdf_list.append(sys.argv[i])

    #send pdfs to convert to text
    pdfToTxt(pdf_list)
